from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
import pandas as pd

import os

from dotenv import load_dotenv
import google.generativeai as genai

from time import sleep


class smart_wait():
    def __init__(self, num_batches:int=10, wait_secs:float=0.0):
        self.num_batches = num_batches
        self.wait_secs = wait_secs
        self.i = 0

    def __iter__(self):
        return self


    def __next__(self):
        self.i += 1
        if self.i%self.num_batches==0:
            sleep(self.wait_secs)

        return


class classification_model():
    def __init__(self, excel_dir:str, num_batches:int=10, wait_secs:float=0.0):
        self.smart_wait = iter(smart_wait(num_batches, wait_secs))

        safety_settings = [
            {
                "category": "HARM_CATEGORY_HARASSMENT",
                "threshold": "BLOCK_NONE"
            },
            {
                "category": "HARM_CATEGORY_HATE_SPEECH",
                "threshold": "BLOCK_NONE"
            },
            {
                "category": "HARM_CATEGORY_SEXUALLY_EXPLICIT",
                "threshold": "BLOCK_NONE"
            },
            {
                "category": "HARM_CATEGORY_DANGEROUS_CONTENT",
                "threshold": "BLOCK_NONE"
            }
        ]
        load_dotenv()
        GOOGLE_API_KEY=os.environ['GOOGLE_API_KEY']
        genai.configure(api_key=GOOGLE_API_KEY)
        self.model = genai.GenerativeModel('gemini-1.0-pro', safety_settings=safety_settings)
        
        self.excel_dir = excel_dir
        self.excel_path = None
        self.book = None
        
        self.sheet = None
        
        self.translation_sheet:Worksheet = None
        self.classification_sheet = None
    
    
    def generate_response(self, prompt:str, original_text):
        next(self.smart_wait)
        response= self.model.generate_content(prompt)
        tries = 0
        while tries<3:
            if response != None:
                if hasattr(response, "text"):
                    print(response.text)
                    return response.text

            print(f"{original_text} failed in {tries+1} iteration to prompt")
            print(f"reason is:\n{response}\n")
            tries+=1
        print(response)
        return response

    def translate_sheet(self):
        rtl_marks = "\u200F"
        # this function translates a person's good and bad comments (one sheet) from hebrew to english
        base_st = "Context: You are a translating chatbot that translates hebrew to english, "
        base_st += "while translating \"מחזור\" to \"year class\" and \"תלפיות\" to \"Talpiot\".\n"
        base_st += "For example: \"אנחנו המחזור הכי טוב בתלפיות\" will be translated to \"We are the best year class in Talpiot\"."
        base_st += "Your response should only include the translation.\n"
        base_st += "User prompt: \""
        
        if "Translation" in self.book.sheetnames:
            self.book.remove(self.book["Translation"])
        self.translation_sheet = self.book.create_sheet("Translation")

        self.translation_sheet['A1'].value, self.translation_sheet['B1'].value = "Original_conserve", "Translation_conserve"
        self.translation_sheet['C1'].value, self.translation_sheet['D1'].value = "Original_improve", "Translation_improve"

        row_fixer = 0
        for i, sentence in enumerate(self.sheet['Q'][1:] + self.sheet['R'][1:]):
            sentence:Cell = sentence
            text = sentence.value
            if sentence.row - row_fixer <= 1:
                row_fixer = 0

            if text is None:
                row_fixer += 1
                continue

            # TODO - add knowing level
            try:
                knowing_level = 'גבוהה' if (int(self.sheet['B'][sentence.row-1].value) > 4) else 'נמוכה'
            except:
                knowing_level = 'נמוכה'

            st = f"{base_st}{text}\""
            
            original_col = 'A' if sentence.column == 17 else 'C'
            translation_col = 'B' if sentence.column == 17 else 'D'

            updated_text = str(text) + f" (רמת היכרות " + knowing_level + f"{rtl_marks}({rtl_marks} "
            self.translation_sheet[f"{original_col}{sentence.row - row_fixer}"].value = updated_text

            self.translation_sheet[f"{translation_col}{sentence.row- row_fixer}"].value = self.generate_response(st, text)
        
        self.book.save(self.excel_path)
    
    
    def set_classifications(self, is_improve:bool, row:int, text:str):
        text = text.lower()
        
        base_col_ascii = ord('I') if is_improve else ord('B')

        if "other" in text:
            self.classification_sheet[f'{chr(base_col_ascii+5)}{row}'].value = "TRUE"
            return

        if "interpersonal skills" in text:
            self.classification_sheet[f'{chr(base_col_ascii)}{row}'].value = "TRUE"
            return

        if "intrapersonal skills" in text:
            self.classification_sheet[f'{chr(base_col_ascii+1)}{row}'].value = "TRUE"
            return

        if "professionalism" in text:
            self.classification_sheet[f'{chr(base_col_ascii+2)}{row}'].value = "TRUE"
            return

        if "conduct" in text:
            self.classification_sheet[f'{chr(base_col_ascii+3)}{row}'].value = "TRUE"
            return

        if "leadership" in text:
            self.classification_sheet[f'{chr(base_col_ascii+4)}{row}'].value = "TRUE"
            return
    
    
    def classify_sheet(self):
        base_st = 'Context: You are a classifying chatbot, that can classify to one category out of the following 6 categories (and only them): '+\
              '1. Interpersonal Skills, 2. Intrapersonal Skills, 3. Professionalism, 4. Conduct, 5. Leadership, 6. Other.'+\
              'So the only allows answers are: "Interpersonal Skills", "Intrapersonal Skills", "Professionalism", "Conduct", "Leadership", or "Other"'+\
              'The "Other" category is user prompts that do not fit very well with any of the categories listed above.'+\
              f'User prompt: "'
              
        if "Classification" in self.book.sheetnames:
            self.book.remove(self.book["Classification"])
        self.classification_sheet = self.book.create_sheet("Classification")

        self.classification_sheet['A1'].value = "Original_conserve"
        self.classification_sheet['B1'].value = "Interpersonal Skills"
        self.classification_sheet['C1'].value = "Intrapersonal Skills"
        self.classification_sheet['D1'].value = "Professionalism"
        self.classification_sheet['E1'].value = "Conduct"
        self.classification_sheet['F1'].value = "Leadership"
        self.classification_sheet['G1'].value = "Other"
        
        self.classification_sheet['H1'].value = "Original_improve"
        self.classification_sheet['I1'].value = "Interpersonal Skills2"
        self.classification_sheet['J1'].value = "Intrapersonal Skills2"
        self.classification_sheet['K1'].value = "Professionalism2"
        self.classification_sheet['L1'].value = "Conduct2"
        self.classification_sheet['M1'].value = "Leadership2"
        self.classification_sheet['N1'].value = "Other2"

        for i in range(2, self.translation_sheet.max_row+1):
            self.classification_sheet[f"A{i}"].value = self.translation_sheet[f"A{i}"].value
            self.classification_sheet[f"H{i}"].value = self.translation_sheet[f"C{i}"].value

            translated_sentence = self.translation_sheet[f"B{i}"].value
            if translated_sentence is not None:
                st = f"{base_st}{translated_sentence}\""
                self.set_classifications(False, i, self.generate_response(st, translated_sentence))

            translated_sentence = self.translation_sheet[f"D{i}"].value
            if translated_sentence is not None:
                st = f"{base_st}{translated_sentence}\""
                self.set_classifications(True, i, self.generate_response(st, translated_sentence))
        
        self.book.save(self.excel_path)
    
    
    def run_classification_for_cadet(self, excel_filename:str, excel_sheet:str=None, run_translation:bool=True, run_classification:bool=True, return_if_classified:bool=True):
        self.excel_path = os.path.join(self.excel_dir, excel_filename)
        self.book = load_workbook(self.excel_path)

        if not return_if_classified:
            if "Classification" in self.book.sheetnames:
                return pd.read_excel(self.excel_path, sheet_name="Classification", engine='openpyxl', header=0)
            else:
                return None

        if excel_sheet is not None:
            self.sheet = self.book[excel_sheet]
        else:
            self.sheet = self.book.active

        if run_translation:
            self.translate_sheet()

        if run_classification:
            self.classify_sheet()

        return pd.read_excel(self.excel_path, sheet_name="Classification", engine='openpyxl', header=0)
        